from openpyxl import load_workbook
from pathlib import Path


WORKBOOK = Path(r"C:\Users\HP\Downloads\Actualiza_tipo_estudiante.xlsx")
OUTPUT = Path(r"C:\Users\HP\OneDrive - Colegio de Profesionales en Informática y Comp\CURSOS ONLINE\PROFE360\actualiza_tipo_estudiante_simple.sql")


def sql_escape(value: str) -> str:
    return value.replace("'", "''")


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        ident = ws.cell(r, 1).value
        tipo = ws.cell(r, 2).value
        if ident is None or str(ident).strip() == "" or tipo is None or str(tipo).strip() == "":
            continue
        ident_str = str(ident).strip()
        tipo_str = str(tipo).strip()
        key = (ident_str, tipo_str)
        if key in seen:
            continue
        seen.add(key)
        rows.append(key)

    lines = []
    lines.append("SET NOCOUNT ON;")
    lines.append("SET XACT_ABORT ON;")
    lines.append("")
    lines.append("DECLARE @Cambios TABLE (")
    lines.append("    Identificacion NVARCHAR(50) NOT NULL,")
    lines.append("    TipoEstudianteDescripcion NVARCHAR(150) NOT NULL")
    lines.append(");")
    lines.append("")
    lines.append("INSERT INTO @Cambios (Identificacion, TipoEstudianteDescripcion) VALUES")
    for i, (ident, tipo) in enumerate(rows):
        sep = "," if i < len(rows) - 1 else ";"
        lines.append(f"    (N'{sql_escape(ident)}', N'{sql_escape(tipo)}'){sep}")
    lines.append("")
    lines.append("BEGIN TRAN;")
    lines.append("")
    lines.append("IF EXISTS (")
    lines.append("    SELECT 1")
    lines.append("    FROM (SELECT DISTINCT TipoEstudianteDescripcion FROM @Cambios) c")
    lines.append("    LEFT JOIN dbo.TipoEstudiante te")
    lines.append("        ON UPPER(LTRIM(RTRIM(te.Descripcion))) = UPPER(LTRIM(RTRIM(c.TipoEstudianteDescripcion)))")
    lines.append("    WHERE te.TipoEstudianteId IS NULL")
    lines.append(")")
    lines.append("BEGIN")
    lines.append("    SELECT DISTINCT c.TipoEstudianteDescripcion AS TipoFaltante")
    lines.append("    FROM (SELECT DISTINCT TipoEstudianteDescripcion FROM @Cambios) c")
    lines.append("    LEFT JOIN dbo.TipoEstudiante te")
    lines.append("        ON UPPER(LTRIM(RTRIM(te.Descripcion))) = UPPER(LTRIM(RTRIM(c.TipoEstudianteDescripcion)))")
    lines.append("    WHERE te.TipoEstudianteId IS NULL;")
    lines.append("")
    lines.append("    RAISERROR('Hay tipos de estudiante que no existen en dbo.TipoEstudiante.', 16, 1);")
    lines.append("    ROLLBACK TRAN;")
    lines.append("    RETURN;")
    lines.append("END;")
    lines.append("")
    lines.append("UPDATE e")
    lines.append("SET e.TipoEstudianteId = te.TipoEstudianteId")
    lines.append("FROM dbo.Estudiante e")
    lines.append("INNER JOIN @Cambios c")
    lines.append("    ON REPLACE(REPLACE(LTRIM(RTRIM(e.Identificacion)), N' ', N''), N'-', N'') = REPLACE(REPLACE(LTRIM(RTRIM(c.Identificacion)), N' ', N''), N'-', N'')")
    lines.append("INNER JOIN dbo.TipoEstudiante te")
    lines.append("    ON UPPER(LTRIM(RTRIM(te.Descripcion))) = UPPER(LTRIM(RTRIM(c.TipoEstudianteDescripcion)));")
    lines.append("")
    lines.append("SELECT")
    lines.append("    @@ROWCOUNT AS FilasActualizadas,")
    lines.append("    (SELECT COUNT(1) FROM @Cambios) AS FilasEnExcel;")
    lines.append("")
    lines.append("COMMIT TRAN;")

    OUTPUT.write_text("\r\n".join(lines), encoding="utf-8-sig")
    print(f"Wrote {OUTPUT} with {len(rows)} rows")


if __name__ == "__main__":
    main()
